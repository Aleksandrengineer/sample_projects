#This code is taking from a specific range of columns their values, delete all n/a and null values
#and then add it to a user specify columns in excel
#But as con, the code doesnt save the formulaes in a new file
#so created file should be used as a endway editing

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

#path = 'C:\\Users\\ASpiridonov\\Desktop\\study\\loader.xlsm'

wb = load_workbook('loader.xlsm', data_only = True)

sheet_obj = wb.active

max_row = sheet_obj.max_column

valuelist = list()

truck_num_list = list()
#this code is for the load cycle time

#this is is getting values from a column range and 
#appends not nulls or n/a to a list
for i in range (1, 707):
	cell_obj = sheet_obj.cell ( column = 23, row = i)
	valuelist.append(cell_obj.value)
values = list()
for value in valuelist:
	if type(value) is int:
		values.append(value)
	else:
		continue

#print(values) #this one checks if i get the list right
k=15
y=24
#this one is takin in consideration the k and y values and 
#insert values in a previous list into the specific column one after one
for i, value in enumerate(values): #using the enumerated i need the position index 'i' 
								   #so to fill the right row
	sheet_obj.cell(column=y, row=k+i, value=value)
	#print (i, value)
wb.save('loadermodified.xlsx')

#this one is for number of bucket for each truck
valuelist.clear()
values.clear()
truck_num_list.clear()
sum = 0
count = 0
row_num = list()
buckets_num = 0
float(buckets_num)
buckets = list()
trucks_buckets = dict()
#this loop is creating a list of the truck number that are presented in the data
for i in range (16, 707):
	cell_obj = sheet_obj.cell (column = 2, row = i)
	if cell_obj.value is not None:
		truck_num_list.append(cell_obj.value)
		truck_num_list = list(set(truck_num_list)) #this is sorting and deleting the duplicates
	else: continue
#print (truck_num_list)
# Here i need to create a dictionary with
for truck_num in truck_num_list:
	sum = 0
	count = 0
	buckets_num = 0
	valuelist.clear()
	row_num.clear()
	#print ('truck number is', truck_num) #this is checking the truck garage number
	for i in range (16, 707):
		cell_obj = sheet_obj.cell (column = 2, row = i)
		if truck_num == cell_obj.value:
			row_num.append(i)
			for m in row_num:
				cell_obj1 = sheet_obj.cell (column = 4, row = m)
				#print (cell_obj1.value) #with the line 67 it is showing us the number of buckets
				#for each truck cycle
				sum = sum+cell_obj1.value
				count = count+1
	buckets_num = sum/count
	print(type(buckets_num))
	round(buckets_num, 0)
	trucks_buckets.update({truck_num : buckets_num})
#print(trucks_buckets)
# these lines is for the assigment of the truck numbers
k=15
y=27
for i, key in enumerate(trucks_buckets):					   
	sheet_obj.cell(column=y, row=k+i, value=key)
wb.save('loadermodified.xlsx')				
# these lines is for assignment of bucket loaded into the trucks, 
k=15
y=29
for i, (key, v) in enumerate(trucks_buckets.items()):					   
	sheet_obj.cell(column=y, row=k+i, value=v)
wb.save('loadermodified.xlsx')
		



wb.save('loadermodified.xlsx')
print('Выполнено, вы можете найти результат в файле с названием loadermodified.xlsx')